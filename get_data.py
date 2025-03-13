# -*- coding: utf-8 -*-
"""
Created on Sat Dec 21 17:13:03 2024

@author: Lenovo
"""

import os
import pandas as pd
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
import openpyxl as op
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.common.keys import Keys
# import requests

# In[0] parameter
excel_path = 'mindmaps.xlsx'
scroll_time = 10 # 根据页面中内容的多少而定。
phone_num = "***" # 手机号（账户名）
password = "***" # 密码
# 如果页面中的内容很多，滚动时无法全部覆盖，可以适当调大滚动次数；
# 如果页面中的内容很少，移动时会在底部停留很久，可以适当调小滚动次数

map_list = [
    ("Home",[
        'Cloakroom',
        'Diningroom',
        'Drawingroom',
        'Garden',
        'GuestBedroom',
        'Kitchen',
        'MasterBedroom',
        'Recreationroom',
        'Storeroom',
        'Studyroom',
        'VirtualReality',
        'Washroom'
        ]),
    ("People",[
        'Augmented Reality',
        'Body Movements',
        'Characters',
        'Daily Activities',
        'Feelings and Expressions',
        'Interpersonal Communication',
        'Life Events',
        'Nature',
        'Occupations',
        'Pastimes and Hobbies',
        'People Categories',
        'The Body'
        ]),
    ("BasicWords",[
        'Adjectives',
        'Animals',
        'Basic Math',
        'Brands',
        'Colors',
        'Countries',
        'Materials',
        'Mixed Reality',
        'Opinions',
        'Opposites',
        'Plants',
        'Time and Units'
        ])
    ]

# In[1] 函数合集
# 创建excel
def create_excel(excelName):
    wb = op.Workbook()
    wb.save(excelName)

# 保留原有数据，新开一个sheet写入数据
def write_toExcel(data, title_name, title_sub_name, fileName):
    wb = op.load_workbook(fileName)
    sheetName = f"{title_sub_name}"
    ws = wb.create_sheet(title=sheetName)
    ws.append([f"{title_name} ---- {title_sub_name}"])
    # ws.cell.font = op.styles.Font(size=24, bold=True)
    for index, value in enumerate(data, start=1):
        index = index+1
        ws.cell(row=index, column=1, value=value)
    wb.save(fileName)
    
# 删除sheet
def delete_sheet(excelName, sheetName):
    wb = op.load_workbook(excelName)
    ws = wb[sheetName]
    wb.remove[ws]
    del wb[sheetName]

# In[2] 爬虫，从幕布抓取数据
# 0.0 创建excel
create_excel(excel_path)

# 1.0 登录幕布
# chromedriver_path = r"C:\Program Files\Google\Chrome\Application\chromedriver.exe"
# driver = webdriver.Chrome(chromedriver_path)
driver = webdriver.Chrome()
url = "https://mubu.com"
driver.get(url) # 打开幕布
original_window = driver.current_window_handle # 获取当前窗口句柄
driver.find_element(By.XPATH, '//button[text()="登录" or text()="进入幕布"]').click() # 点击“登录”按钮
time.sleep(0.5) # 等待网页刷新
driver.find_element(By.XPATH, '//button[text()="切换到密码登录"]').click() # 切换为密码登录模式
time.sleep(0.5) # 等待网页刷新
driver.find_element(By.XPATH, '//input[@placeholder="输入手机号"]').send_keys(phone_num) # 输入账户
driver.find_element(By.XPATH, '//input[@placeholder="密码"]').send_keys(password) # 输入密码

driver.find_element(By.XPATH, '//button[@class="Buttonstyle-gnoclh-1 dVkztK sc-bOCgKg emIcI sc-bOCgKg emIcI"]').click() # 点击登录
time.sleep(2) # 等待网页刷新

# In[test]
# 2.0 进入幕布文件夹
all_windows = driver.window_handles # 获取所有窗口句柄
for window in all_windows:
    if window !=original_window:
        driver.switch_to.window(window) # 切换到新窗口
        break

# 3.0 获取跳转链接
# 3.1 打开词典文件夹
title_ori_name = "词典"
# title_ori_name = "英语"

html_content = driver.page_source # 获取html内容
soup = BeautifulSoup(html_content, 'html.parser') # 解析html内容
scroll_view_div = soup.find('div', id='js-documents-tree-scroll-view') # 获取跳转目录
title_div = scroll_view_div.find('div', title=title_ori_name) # 通过标题定位
if title_div:
    # previous_sibling = title_div.find_previous_siblings('div')[2]
    id_div = title_div.parent.parent
    # print(id_div)
    select_ori_id = id_div.get('data-selectid')
    # print(select_ori_id)

# In[test]
# 3.2 打开一级文件夹，如people
for jj in range(1,len(map_list)+1):
    map_num = jj-1
    
    if map_num:
        print(f"\n正在重新打开{title_ori_name}")
    else:
        print(f"\n正在打开{title_ori_name}")
    url = f"https://mubu.com/app/folders/home/{select_ori_id}" # 获取链接
    driver.get(url) # 打开链接
    time.sleep(3)
    
    title_name = map_list[map_num][0]
    # title_name = "test"
    html_content = driver.page_source # 获取html内容
    soup = BeautifulSoup(html_content, 'html.parser') # 解析html内容
    scroll_view_div = soup.find('div', id='js-documents-tree-scroll-view') # 获取跳转目录
    title_div = scroll_view_div.find('div', title=title_name) # 通过标题定位
    if title_div:
        id_div = title_div.parent.parent
        select_id = id_div.get('data-selectid') # 获取跳转id
    
    for ii in range(1,len(map_list[map_num][1])+1):
        index = ii-1
        
        if index:
            print(f"\n正在重新打开{title_name}")
        else:
            print(f"\n正在打开{title_name}")
        url = f"https://mubu.com/app/folders/home/{select_id}" # 获取链接
        driver.get(url) # 打开链接
        time.sleep(3)
        
        # 3.3 操作二级文件夹，如people下的the body
        # 3.3.1 打开网页
        title_sub_name = map_list[map_num][1][index]
        # title_sub_name = "haha"
        
        html_content = driver.page_source # 获取html内容
        soup = BeautifulSoup(html_content, 'html.parser') # 解析html内容
        scroll_view_div = soup.find('div', id='js-documents-tree-scroll-view') # 获取跳转目录
        title_div = scroll_view_div.find('div', title=title_sub_name) # 通过标题定位
        if title_div:
            id_div = title_div.parent.parent
            select_sub_id = id_div.get('data-selectid') # 获取跳转id
        
        print(f"\n正在打开{title_name}下的{title_sub_name}")
        url = f"https://mubu.com/app/edit/home/{select_sub_id}" # 获取链接
        driver.get(url) # 打开链接
        time.sleep(3)
        
        # 3.3.2 滚动页面，加载全部内容
        # document.body.scrrollHeight
        # driver.execute_script("window.scrollTo(0, 1500);")
        # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # ans = driver.execute_script("return document.body.scrollHeight")
        # driver.execute_script("window.scrollTop +=1000;")
        
        # 滚动scroll组件：选中scroll组件后，再移动
        scroll_div = driver.find_element(By.ID, "js-doc-wrap")
        for _ in range(scroll_time):
            driver.execute_script("arguments[0].scrollBy(0,200);", scroll_div)
            time.sleep(0.5)

      # test starts ----------------------------------
        # 从上向下滚动
        #   body = driver.find_element(By.TAG_NAME, 'body')
        # for _ in range(2):
        #     ActionChains(driver).move_to_element(body).send_keys(Keys.PAGE_DOWN).perform()
        #     time.sleep(0.5)
        # time.sleep(3)
        
        # 从底部向上滚动 #
        # target = driver.find_element(By.XPATH, '//footer[@class="footer"]/div')
        # for _ in range(2):
        #     ActionChains(driver).move_to_element(target).send_keys(Keys.PAGE_UP).perform()
        #     time.sleep(0.5)
        
        # 直接滚动到特定位置
        # target = driver.find_element(By.XPATH, '//footer[@class="footer"]/div')
        # driver.execute_script("arguments[0].scrollIntoView();", target) # 移动至页尾的地方
        # --- ---#   
        
        # 缩放页面
        # zoom_level = 1.5
        # driver.execute_script("document.body.style.zoom='{}%';".format(zoom_level*100))
        
        # 使用requests.post向服务器发起访问请求
        # url = "https://api2.mubu.com/v3/api/list/star_relation/get"
        # # headers = {
        # #     # 'Referer':"https://mubu.com/",
        # #     # "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        # #     "Content-Type":"application/json;charset=UTF-8",
        # #     "jwt-token": "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJhcHAiOiJtdWJ1Iiwic3ViIjoiMTM0NTkwMzYiLCJsb2dpblR5cGUiOiJtb2JpbGUiLCJleHAiOjE3MzgxNTk0MzUsImlhdCI6MTczNTU2NzQzNX0.VSQPRtpljkLZRI_VXompOsnwREHekSfxtJ5SL_Zx2fWBPqyO0Mv-akkYc4XbP2qoA0e-CtoLGA1NJV9JSNNFYA"
        # #     }
        # headers = {
        #     "X-Reg-Entrance":"https://mubu.com/app/edit/home/69HHPHenqw0",
        #     "X-Request-Id": "9f353f69-db95-4dee-808d-a7f7260667cf",
        #     "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        #     "Content-Type":"application/json;charset=UTF-8",
        #     "jwt-token": "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJhcHAiOiJtdWJ1Iiwic3ViIjoiMTM0NTkwMzYiLCJsb2dpblR5cGUiOiJtb2JpbGUiLCJleHAiOjE3MzgxNTk0MzUsImlhdCI6MTczNTU2NzQzNX0.VSQPRtpljkLZRI_VXompOsnwREHekSfxtJ5SL_Zx2fWBPqyO0Mv-akkYc4XbP2qoA0e-CtoLGA1NJV9JSNNFYA"
        #     }
        # # params = {
        # #     "x-reg-entrance":"https://mubu.com/app/edit/home/69HHPHenqw0",
        # #     "x-request-id":"3e09f23a-2d85-41ab-a929-bbeb575116cf"
        # #     }
        # payload = {
        #     "docId": "69HHPHenqw0",
        #     "password": "",
        #     "isFromDocDir": 1
        #     }
        # # response = requests.get(url, headers = headers, params = params)
        # response = requests.post(url, data=payload, headers=headers)
        # data=response.json()
        # print(data)
        # print(f"下面是cookies信息：{data}")
      # test ends ----------------------------------
        
        # 3.3.3 获取页面内容
        html_content = driver.page_source # 获取html内容
        soup = BeautifulSoup(html_content, 'html.parser') # 解析html内容
        # 获取思维导图
        div_element = soup.find('div', class_='outliner-tree') # 获取思维导图内容
        # span_contents = [span.get_text().strip() for span in div_element.find_all('span') if not span.get('class')] # 如果需要去掉思维导图中的分类名，而只保留最低一级的内容，则使用这个代码
        span_contents = [span.get_text().strip() for span in div_element.find_all('span') if span.get_text().strip()] # 截取span中的有效内容，即思维导图，一个一行值
        # 获取思维导图概述（有多少条主题，多少字）
        footer = soup.find('footer').find('div')
        len_element = footer.get_text()
        print(f"目前载入了{len(span_contents)}条数据")
        print(len_element)
        
        # 3.3.4 写入excel中
        write_toExcel(span_contents, title_name, title_sub_name, excel_path)
        print(f"————————{title_name}下的{title_sub_name}对应的mindmap已写入excel中")
   
# delete_sheet(excel_path, "Sheet")
driver.quit()

# In[3] 比对数据
excel_path = 'mindmaps--20250101.xlsx'
flag = 1
for map_num in range(0,len(map_list)):
    map_name = map_list[map_num][0]
    for index in range(0,len(map_list[map_num][1])):
        curr_map = map_list[map_num][1][index]
        
        # set the destination path
        sheet_name = f"{curr_map}"
        folder_path = f"Project_DailyLife_{map_name}_Prepare\{map_name}\Images_{map_name}\{curr_map}"
    
        # read excel
        df = pd.read_excel(excel_path, sheet_name=sheet_name, usecols=[0])
        words = df.iloc[:,0].dropna().str.lower().tolist()
        words = [word.rstrip() for word in words] #去除字符末尾的空格
        
        # exam and record
        unmatched_words = []
        unpaired_pics = []
        
        image_files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
        image_names = [os.path.splitext(f)[0].lower() for f in image_files]
        
        for word in words:
            image_name = f"{word}.jpg"
            image_path = os.path.join(folder_path, image_name)
            if not os.path.exists(image_path):
                if re.search(r'\d',word): # 把带数字的字符去掉（带数字的是分类名）
                    continue
                # if any(char.isdigit for char in word):
                #     continue
                # for char in word:
                #     print(f"{char}:::{char.isdigit}")
                # print(f"end{word}")
                unmatched_words.append(word)
                
        # print(f"\n对于{sheet_name}:\n")
        
        # report unmatched words
        if unmatched_words:
            print(f"\n对于{sheet_name},以下单词没有对应的图片：")
            # print("以下单词没有对应的图片：")
            for word in unmatched_words:
                flag = 0
                print(word)
        # else:
        #     print("所有单词均有对应图片\n")
            
        # report unpaired pictures
        unpaired_pics = [name for name in image_names if name not in words]
        if unpaired_pics:
            print(f"\n对于{sheet_name},以下图片没有在思维导图中出现：")
            # print("以下图片没有在思维导图中出现")
            for image in unpaired_pics:
                flag = 0
                print(image)
        # else:
        #     print("所有图片都在思维导图中")
if flag:
    print("\n芜湖！恭喜！mindmap和图片完成了同步！")
# for mm in range(0,len(ans1)):
#     if(ans1[mm]!=ans2[mm]):
#         print(mm)
#         print(ans1[mm])
#         print(ans2[mm])
                    
                
